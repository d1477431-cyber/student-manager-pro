import csv
import json
import io
from decimal import Decimal
from datetime import datetime, date, timedelta

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Sum, Count, Avg
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from .forms import StudentForm, ChangePasswordForm, RegisterForm
from .logic import calculate_average, get_appreciation, parse_notes, sanitize_cell, compute_statut_paiement
from .realtime import notify_user
from .models import (
    Student, CustomUser, Log, Payment, Echeance, Absence, Cours, Message, Note,
    Annonce, Notification, ALLOWED_ATTACHMENT_EXTENSIONS,
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


def health_check(request):
    """Vue publique pour le health check de Railway"""
    return HttpResponse("OK", status=200)


def _authenticate_fallback(username, password):
    if not username or not password:
        return None
    return authenticate(username=username, password=password)


def login_view(request):
    if request.user.is_authenticated:
        return redirect('index')
    
    error_message = None
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        
        try:
            custom_user = CustomUser.objects.get(username=username)
            
            # La méthode `can_login` centralise les vérifications
            can_login, reason = custom_user.can_login()
            if not can_login:
                if "verrouillé" in reason:
                    Log.objects.create(username=username, event='LOGIN_BLOCKED', details=reason)
                else:
                    Log.objects.create(username=username, event='LOGIN_FAILURE', details=reason)
                return render(request, 'registration/login.html', {'error': reason})

        except CustomUser.DoesNotExist:
            # Si l'utilisateur n'existe pas, on ne donne pas d'info spécifique
            # pour des raisons de sécurité, mais on logue l'échec.
            Log.objects.create(username=username, event='LOGIN_FAILURE', details="Utilisateur inconnu")
            error_message = "❌ Identifiants incorrects."
            return render(request, 'registration/login.html', {'error': error_message})
        
        user = _authenticate_fallback(username, password)
        
        if user is not None:
            try:
                custom_user = CustomUser.objects.get(username=username)
                custom_user.reset_login_attempts()
            except CustomUser.DoesNotExist:
                pass
            
            login(request, user)
            Log.objects.create(username=username, event='LOGIN_SUCCESS',
                               details=f"Connexion réussie")
            messages.success(request, 'Connexion réussie.')
            
            # Forcer le changement de mot de passe si must_change_password est True
            try:
                cu = CustomUser.objects.get(username=username)
                if cu.must_change_password:
                    return redirect('change_password')
            except CustomUser.DoesNotExist:
                pass
            
            return redirect('index')
        else:
            try:
                custom_user = CustomUser.objects.get(username=username)
                custom_user.increment_login_attempts()
                remaining_attempts = 5 - custom_user.login_attempts
                if remaining_attempts > 0:
                    error_message = f"❌ Identifiants incorrects. Il vous reste {remaining_attempts} tentative(s)."
                else:
                    error_message = "🔒 Compte verrouillé pour 15 minutes suite à trop de tentatives."
            except CustomUser.DoesNotExist:
                error_message = "❌ Identifiants incorrects."
            
            Log.objects.create(username=username, event='LOGIN_FAILURE',
                               details=f"Tentative échouée")
    
    return render(request, 'registration/login.html', {'error': error_message})


def register_view(request):
    """Auto-inscription d'un compte Professeur, en attente d'approbation par un admin."""
    if request.user.is_authenticated:
        return redirect('index')

    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = CustomUser.objects.create(
                username=form.cleaned_data['username'],
                first_name=form.cleaned_data.get('first_name', ''),
                last_name=form.cleaned_data.get('last_name', ''),
                email=form.cleaned_data.get('email', ''),
                password=make_password(form.cleaned_data['password']),
                role='Professeur',
                status='pending',
                must_change_password=False,
            )
            Log.objects.create(
                username=user.username, event='REGISTER_REQUEST',
                details="Demande d'inscription Professeur (en attente d'approbation)"
            )
            messages.success(
                request,
                "✅ Votre demande d'inscription a été envoyée. Un administrateur doit "
                "approuver votre compte avant que vous puissiez vous connecter."
            )
            return redirect('login')
    else:
        form = RegisterForm()

    return render(request, 'registration/register.html', {'form': form})


@login_required
def change_password_view(request):
    """Vue pour forcer le changement de mot de passe"""
    custom_user = request.user  # déjà un CustomUser (AUTH_USER_MODEL)

    if request.method == 'POST':
        form = ChangePasswordForm(request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['new_password']
            custom_user.set_password(new_password)
            custom_user.must_change_password = False
            custom_user.password_changed_at = timezone.now()
            custom_user.save()
            
            update_session_auth_hash(request, custom_user)
            
            Log.objects.create(username=request.user.username, event='PASSWORD_CHANGE',
                               details="Mot de passe changé (première connexion)")
            messages.success(request, '✅ Mot de passe modifié avec succès.')
            return redirect('index')
    else:
        form = ChangePasswordForm()
    
    return render(request, 'registration/change_password.html', {
        'form': form,
        'username': request.user.username,
    })


def logout_view(request):
    Log.objects.create(username=request.user.username, event='LOGOUT',
                       details="Déconnexion")
    logout(request)
    messages.success(request, 'Déconnexion réussie.')
    return redirect('login')


# ===== DECORATEUR DE PERMISSIONS =====
def permission_required(perm):
    """Décorateur pour vérifier les permissions par rôle"""
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            # request.user est déjà un CustomUser (AUTH_USER_MODEL) : pas besoin
            # de le requêter à nouveau. Ce décorateur protège ~15 vues, donc
            # cette requête en moins compte sur une base distante (Supabase).
            if request.user.is_authenticated and not request.user.has_permission(perm):
                messages.error(request, "⛔ Vous n'avez pas les droits nécessaires pour cette action.")
                Log.objects.create(username=request.user.username, event='PERMISSION_DENIED',
                                   details=f"Tentative d'accès à {perm}")
                return redirect('index')
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


@login_required
@permission_required('can_export_data')
def export_students_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="etudiants.csv"'
    writer = csv.writer(response)
    writer.writerow(['Matricule', 'Nom', 'Prénom', 'Filière', 'Niveau', 'Statut', 'Notes', 'Moyenne', 'Appréciation',
                     'Frais Scolarité', 'Total Payé', 'Solde', 'Statut Paiement'])
    students = Student.objects.all().order_by('-matricule').prefetch_related('notes')
    for s in students:
        notes_list = [float(n.valeur) for n in s.notes.all()]
        notes_str = ', '.join(str(round(n, 2)) for n in notes_list)
        moyenne = calculate_average(notes_list)
        writer.writerow([sanitize_cell(v) for v in [
            s.matricule, s.nom, s.prenom, s.filiere or '-', s.niveau or '-', s.statut,
            notes_str, round(moyenne, 2), get_appreciation(moyenne),
            float(s.frais_scolarite), float(s.total_paye()), float(s.solde_restant()), s.statut_paiement()
        ]])
    return response


@login_required
@permission_required('can_export_data')
def export_students_excel(request):
    if not OPENPYXL_AVAILABLE:
        messages.error(request, 'openpyxl n\'est pas installé.')
        return redirect('index')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Étudiants"
    
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4F46E5', end_color='7C3AED', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ['Matricule', 'Nom', 'Prénom', 'Filière', 'Niveau', 'Statut', 'Moyenne', 'Appréciation',
               'Frais Scolarité', 'Total Payé', 'Solde', 'Statut Paiement']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    students = Student.objects.all().order_by('-matricule').prefetch_related('notes')
    for row_idx, s in enumerate(students, 2):
        notes_list = [float(n.valeur) for n in s.notes.all()]
        moyenne = calculate_average(notes_list)
        data = [sanitize_cell(v) for v in [
            s.matricule, s.nom, s.prenom, s.filiere or '-', s.niveau or '-', s.statut,
            round(moyenne, 2), get_appreciation(moyenne),
            float(s.frais_scolarite), float(s.total_paye()), float(s.solde_restant()), s.statut_paiement()
        ]]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="etudiants.xlsx"'
    wb.save(response)
    return response


@login_required
def index(request):
    query = request.GET.get('q', '').strip()
    statut_filter = request.GET.get('statut', '').strip()
    paiement_filter = request.GET.get('paiement', '').strip()
    filiere_filter = request.GET.get('filiere', '').strip()
    niveau_filter = request.GET.get('niveau', '').strip()
    
    # Vérifier les permissions d'affichage selon le rôle
    # request.user est déjà un CustomUser (AUTH_USER_MODEL) : pas besoin de le
    # requêter à nouveau (une requête de moins sur cette page à fort trafic)
    cu = request.user
    can_view_finances = cu.has_permission('can_view_finances')
    can_add_student = cu.has_permission('can_add_student')
    can_export_data = cu.has_permission('can_export_data')
    can_generate_documents = cu.has_permission('can_generate_documents')
    can_delete_student = cu.has_permission('can_delete_student')

    # Annoté avec le total payé (une requête agrégée) plutôt que d'appeler
    # total_paye()/statut_paiement() en boucle plus bas (N+1 queries)
    students = Student.with_totals(Student.objects.all().prefetch_related('notes'))

    if query:
        search_filter = Q(nom__icontains=query) | Q(prenom__icontains=query)
        if query.isdigit():
            search_filter |= Q(matricule=int(query))
        students = students.filter(search_filter)
    if statut_filter:
        students = students.filter(statut=statut_filter)
    if filiere_filter:
        students = students.filter(filiere__iexact=filiere_filter)
    if niveau_filter:
        students = students.filter(niveau__iexact=niveau_filter)
    if paiement_filter:
        wanted = {'a_jour': 'À jour', 'partiel': 'Partiel', 'en_retard': 'En retard'}.get(paiement_filter)
        if wanted:
            students = [s for s in students if compute_statut_paiement(s.total_paye_annot, s.frais_scolarite) == wanted]

    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(students, 25)  # 25 étudiants par page
    try:
        students_page = paginator.page(page)
    except PageNotAnInteger:
        students_page = paginator.page(1)
    except EmptyPage:
        students_page = paginator.page(paginator.num_pages)

    student_data = []
    for s in students_page:
        notes_qs = s.notes.all()
        notes_list = [float(n.valeur) for n in notes_qs]
        moyenne = calculate_average(notes_list)
        student_data.append({
            'student': s,
            'moyenne': moyenne,
            'appreciation': get_appreciation(moyenne),
            'notes_list': notes_list,
            'total_paye': s.total_paye_annot if can_view_finances else None,
            'solde': (s.frais_scolarite - s.total_paye_annot) if can_view_finances else None,
            'statut_paiement': compute_statut_paiement(s.total_paye_annot, s.frais_scolarite) if can_view_finances else None,
            'photo_url': s.photo_url(),
        })
    
    # Statistiques calculées sur le total (pas seulement la page)
    total = students.count() if hasattr(students, 'count') else len(students)
    
    moyennes = [item['moyenne'] for item in student_data]
    moyenne_generale = sum(moyennes) / len(moyennes) if moyennes else 0.0
    meilleurs_etudiants = sorted(student_data, key=lambda item: item['moyenne'], reverse=True)[:3]
    
    actifs = Student.objects.filter(statut='Actif').count()
    suspendus = Student.objects.filter(statut='Suspendu').count()
    
    # Statistiques financières (uniquement si permission)
    total_scolarite = total_encaisse = total_dettes = a_jour = en_retard = partiel = 0
    alertes = []
    chart_paiements = chart_impayes = None
    
    if can_view_finances:
        # Annotation en une seule requête (total payé par étudiant) plutôt que
        # d'appeler total_paye()/statut_paiement() en boucle (N+1 queries)
        all_students = list(Student.with_totals())
        total_scolarite = sum(float(s.frais_scolarite) for s in all_students)
        total_encaisse = sum(float(s.total_paye_annot) for s in all_students)
        total_dettes = total_scolarite - total_encaisse
        statuts = {
            s.matricule: compute_statut_paiement(s.total_paye_annot, s.frais_scolarite)
            for s in all_students
        }
        a_jour = sum(1 for v in statuts.values() if v == 'À jour')
        en_retard = sum(1 for v in statuts.values() if v == 'En retard')
        partiel = sum(1 for v in statuts.values() if v == 'Partiel')
        alertes = [s for s in all_students if statuts[s.matricule] == 'En retard'][:5]
        
        # Graphiques : données préparées ici, rendues en Chart.js côté navigateur
        # (le rendu serveur via Matplotlib était coûteux en CPU, pénible sur
        # l'hébergement gratuit — voir chart_paiements/chart_impayes ci-dessous)
        raw_paiements = Payment.objects.filter(
            date_paiement__gte=timezone.now() - timedelta(days=365)
        ).values_list('date_paiement', 'montant').order_by('date_paiement')

        mois_dict = {}
        for dp, montant in raw_paiements:
            if dp:
                key = dp.strftime('%Y-%m') if hasattr(dp, 'strftime') else str(dp)[:7]
                mois_dict[key] = mois_dict.get(key, 0) + float(montant)

        paiements_list = sorted(mois_dict.items())
        if paiements_list:
            chart_paiements = {
                'labels': [m[-2:] for m, _ in paiements_list],
                'values': [round(v, 2) for _, v in paiements_list],
            }

        # Taux d'impayés par filière, calculé en une seule requête agrégée
        # (au lieu d'une requête total_paye() par étudiant et par filière)
        filieres_totals = {}
        for s in Student.with_totals().exclude(filiere__isnull=True).exclude(filiere=''):
            entry = filieres_totals.setdefault(s.filiere, {'du': Decimal('0.00'), 'paye': Decimal('0.00')})
            entry['du'] += s.frais_scolarite
            entry['paye'] += s.total_paye_annot

        noms_filieres, taux_impayes = [], []
        for filiere, totaux in filieres_totals.items():
            if totaux['du'] > 0:
                noms_filieres.append(filiere)
                taux_impayes.append(round(float((totaux['du'] - totaux['paye']) / totaux['du'] * 100), 1))

        if noms_filieres:
            chart_impayes = {'labels': noms_filieres, 'values': taux_impayes}
    
    filieres = Student.objects.values_list('filiere', flat=True).distinct().exclude(filiere__isnull=True).exclude(filiere='')
    niveaux = Student.objects.values_list('niveau', flat=True).distinct().exclude(niveau__isnull=True).exclude(niveau='')
    
    return render(request, 'students/index.html', {
        'students': students_page,
        'student_data': student_data,
        'moyenne_generale': moyenne_generale,
        'total_etudiants': total,
        'actifs': actifs,
        'suspendus': suspendus,
        'meilleure_moyenne': max(moyennes) if moyennes else 0,
        'meilleurs_etudiants': meilleurs_etudiants,
        'query': query, 'statut_filter': statut_filter,
        'paiement_filter': paiement_filter, 'filiere_filter': filiere_filter, 'niveau_filter': niveau_filter,
        'filieres': filieres, 'niveaux': niveaux,
        'can_view_finances': can_view_finances,
        'can_add_student': can_add_student,
        'can_export_data': can_export_data,
        'can_generate_documents': can_generate_documents,
        'can_delete_student': can_delete_student,
        'total_scolarite': total_scolarite, 'total_encaisse': total_encaisse, 'total_dettes': total_dettes,
        'a_jour': a_jour, 'en_retard': en_retard, 'partiel': partiel,
        'alertes': alertes,
        'chart_paiements': chart_paiements, 'chart_impayes': chart_impayes,
    })


@login_required
@permission_required('can_add_student')
def add_student(request):
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            if Student.objects.filter(matricule=form.cleaned_data['matricule']).exists():
                messages.error(request, 'Ce matricule existe déjà.')
            else:
                notes_values = parse_notes(form.cleaned_data.get('note_str', ''))
                student = Student(
                    matricule=form.cleaned_data['matricule'],
                    nom=form.cleaned_data['nom'], prenom=form.cleaned_data['prenom'],
                    age=form.cleaned_data['age'], date_ajout=timezone.now(),
                    email=form.cleaned_data['email'], telephone=form.cleaned_data['telephone'],
                    statut=form.cleaned_data['statut'], filiere=form.cleaned_data['filiere'],
                    niveau=form.cleaned_data['niveau'],
                    frais_scolarite=form.cleaned_data['frais_scolarite'] or 0.0,
                )
                if 'photo' in request.FILES:
                    student.photo = request.FILES['photo']
                student.save()
                
                # Créer les notes via le modèle dédié
                for val in notes_values:
                    Note.objects.create(student=student, valeur=val)
                
                frais = float(student.frais_scolarite)
                montant_tranche = round(frais / 3, -2)
                today = timezone.now().date()
                for i in range(3):
                    Echeance.objects.create(
                        student=student, numero_tranche=i + 1,
                        montant=montant_tranche if i < 2 else frais - (montant_tranche * 2),
                        date_echeance=today + timedelta(days=(i + 1) * 90),
                    )
                
                Log.objects.create(username=request.user.username, event='STUDENT_ADD',
                                   details=f"Ajout: {form.cleaned_data['nom']} {form.cleaned_data['prenom']} ({form.cleaned_data['matricule']})")
                messages.success(request, 'Étudiant ajouté avec succès.')
                return redirect('index')
        else:
            messages.error(request, 'Veuillez corriger le formulaire.')
    else:
        form = StudentForm()
    return render(request, 'students/add_student.html', {'form': form})


@login_required
def edit_student(request, matricule):
    student = get_object_or_404(Student, matricule=matricule)
    if request.method == 'POST':
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            notes_values = parse_notes(form.cleaned_data.get('note_str', ''))
            student.nom = form.cleaned_data['nom']
            student.prenom = form.cleaned_data['prenom']
            student.age = form.cleaned_data['age']
            # Mettre à jour les notes via le modèle dédié
            student.notes.all().delete()
            for val in notes_values:
                Note.objects.create(student=student, valeur=val)
            student.email = form.cleaned_data['email']
            student.telephone = form.cleaned_data['telephone']
            student.statut = form.cleaned_data['statut']
            student.filiere = form.cleaned_data['filiere']
            student.niveau = form.cleaned_data['niveau']
            student.frais_scolarite = form.cleaned_data['frais_scolarite'] or 0.0
            if 'photo' in request.FILES:
                student.photo = request.FILES['photo']
            student.save()
            Log.objects.create(username=request.user.username, event='STUDENT_UPDATE',
                               details=f"Modification: {form.cleaned_data['nom']} {form.cleaned_data['prenom']} ({matricule})")
            messages.success(request, 'Étudiant modifié.')
            return redirect('index')
    else:
        notes_list = [float(n.valeur) for n in student.notes.all()]
        initial_notes = ', '.join(str(round(n, 2)) for n in notes_list)
        form = StudentForm(initial={
            'matricule': student.matricule, 'nom': student.nom, 'prenom': student.prenom,
            'age': student.age, 'note_str': initial_notes, 'filiere': student.filiere,
            'niveau': student.niveau, 'email': student.email, 'telephone': student.telephone,
            'statut': student.statut, 'frais_scolarite': student.frais_scolarite,
        })
        form.fields['matricule'].disabled = True
    return render(request, 'students/edit_student.html', {'form': form, 'student': student})


@login_required
@permission_required('can_delete_student')
def delete_student(request, matricule):
    if request.method == 'POST':
        student = get_object_or_404(Student, matricule=matricule)
        nom_complet = f"{student.nom} {student.prenom}"
        student.delete()
        Log.objects.create(username=request.user.username, event='STUDENT_DELETE',
                           details=f"Suppression: {nom_complet} ({matricule})")
        messages.success(request, 'Étudiant supprimé.')
    return redirect('index')


# ==================== PAIEMENTS ====================

@login_required
@permission_required('can_view_finances')
def paiements_view(request):
    paiements = Payment.objects.all().order_by('-date_paiement')[:50]
    students = Student.objects.all()
    echeances = Echeance.objects.filter(statut__in=['en_attente', 'en_retard']).order_by('date_echeance')[:20]
    
    total_scolarite = sum(float(s.frais_scolarite) for s in students)
    total_encaisse = sum(float(p.montant) for p in Payment.objects.all())
    total_dettes = total_scolarite - total_encaisse
    alertes_echeances = [e for e in Echeance.objects.all() if e.est_en_retard()][:10]
    
    return render(request, 'students/paiements.html', {
        'students': students, 'paiements': paiements, 'echeances': echeances,
        'alertes_echeances': alertes_echeances,
        'total_scolarite': total_scolarite, 'total_encaisse': total_encaisse, 'total_dettes': total_dettes,
    })


@login_required
@permission_required('can_view_finances')
def enregistrer_paiement(request):
    if request.method == 'POST':
        matricule = request.POST.get('matricule', '').strip()
        montant_str = request.POST.get('montant', '').strip()
        type_paiement = request.POST.get('type_paiement', 'Espèces')
        
        if not matricule or not montant_str:
            messages.error(request, 'Veuillez remplir tous les champs.')
            return redirect('paiements')
        
        try:
            montant = float(montant_str)
            student = get_object_or_404(Student, matricule=matricule)
            
            payment = Payment.objects.create(
                student=student, montant=montant, date_paiement=timezone.now(), type_paiement=type_paiement,
            )
            
            echeances = Echeance.objects.filter(student=student, statut='en_attente').order_by('numero_tranche')
            reste = montant
            for echeance in echeances:
                if reste <= 0:
                    break
                montant_echeance = float(echeance.montant)
                if reste >= montant_echeance:
                    echeance.statut = 'payee'
                    echeance.date_paiement = timezone.now()
                    echeance.save()
                    reste -= montant_echeance
                else:
                    echeance.montant = Decimal(str(montant_echeance - reste))
                    echeance.save()
                    reste = 0
            
            Log.objects.create(username=request.user.username, event='PAYMENT_ADD',
                               details=f"#{payment.numero_recu} - {montant:,.0f} FCFA - {student.nom} {student.prenom}")
            messages.success(request, f'Paiement #{payment.numero_recu} de {montant:,.0f} FCFA enregistré.')
        except ValueError:
            messages.error(request, 'Montant invalide.')
    return redirect('paiements')


@login_required
@permission_required('can_view_finances')
def generer_recu_pdf(request, payment_id):
    if not REPORTLAB_AVAILABLE:
        messages.error(request, 'ReportLab n\'est pas installé.')
        return redirect('paiements')
    
    payment = get_object_or_404(Payment, id=payment_id)
    student = payment.student
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="recu_{payment.numero_recu}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], alignment=1, spaceAfter=20,
                                  textColor=colors.HexColor('#4f46e5'))
    elements.append(Paragraph("<b>STUDENT MANAGER PRO</b>", title_style))
    elements.append(Paragraph(f"<b>REÇU DE PAIEMENT N° {payment.numero_recu}</b>", styles['Heading2']))
    elements.append(Spacer(1, 20))
    
    data = [
        ['DÉSIGNATION', 'DÉTAILS'],
        ['Date', payment.date_paiement.strftime('%d/%m/%Y %H:%M')],
        ['Étudiant', f"{student.nom} {student.prenom}"],
        ['Matricule', student.matricule],
        ['Filière', student.filiere or '-'],
        ['Niveau', student.niveau or '-'],
        ['Mode', payment.type_paiement],
        ['Montant Versé', f"{float(payment.montant):,.0f} FCFA"],
        ['', ''],
        ['Total Frais', f"{float(student.frais_scolarite):,.0f} FCFA"],
        ['Total versé', f"{float(student.total_paye()):,.0f} FCFA"],
        ['RESTE À PAYER', f"<b>{float(student.solde_restant()):,.0f} FCFA</b>"],
    ]
    
    table = Table(data, colWidths=[6*cm, 10*cm])
    table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('PADDING', (0, 0), (-1, -1), 10),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (0, 8), (-1, -1), colors.HexColor('#f8fafc')),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("Signature : _______________________", styles['Normal']))
    elements.append(Paragraph(f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}", styles['Normal']))
    
    doc.build(elements)
    return response


# ==================== ÉCHÉANCES ====================

@login_required
@permission_required('can_view_finances')
def generer_echeances(request, matricule):
    student = get_object_or_404(Student, matricule=matricule)
    if request.method == 'POST':
        nb_tranches = int(request.POST.get('nb_tranches', 3))
        frais = float(student.frais_scolarite)
        montant_tranche = round(frais / nb_tranches, -2)
        today = timezone.now().date()
        
        Echeance.objects.filter(student=student).delete()
        
        for i in range(nb_tranches):
            montant = montant_tranche if i < nb_tranches - 1 else frais - (montant_tranche * (nb_tranches - 1))
            Echeance.objects.create(
                student=student, numero_tranche=i + 1, montant=montant,
                date_echeance=today + timedelta(days=(i + 1) * 90),
            )
        
        messages.success(request, f'{nb_tranches} échéances générées.')
    return redirect('paiements')


# ==================== LOGS ====================

@login_required
@permission_required('can_view_logs')
def logs_view(request):
    logs = Log.objects.all()[:100]
    return render(request, 'students/logs.html', {'logs': logs})


# ==================== UTILISATEURS ====================

@login_required
@permission_required('can_manage_users')
def utilisateurs_view(request):
    users = CustomUser.objects.exclude(status='pending').order_by('username')
    pending_users = CustomUser.objects.filter(status='pending').order_by('-date_joined')
    return render(request, 'students/utilisateurs.html', {
        'users': users,
        'pending_users': pending_users,
    })


@login_required
@permission_required('can_manage_users')
def approuver_utilisateur(request, user_id):
    """Approuve une demande d'inscription en attente : active le compte."""
    if request.method == 'POST':
        target = get_object_or_404(CustomUser, id=user_id, status='pending')
        target.status = 'active'
        target.save(update_fields=['status'])
        Log.objects.create(username=request.user.username, event='USER_APPROVE',
                           details=f"Compte approuvé : {target.username}")
        Notification.objects.create(
            destinataire=target, type='systeme', titre='Compte approuvé',
            message="Votre compte a été approuvé. Vous pouvez maintenant vous connecter.",
        )
        messages.success(request, f"✅ Compte de {target.username} approuvé.")
    return redirect('utilisateurs')


@login_required
@permission_required('can_manage_users')
def rejeter_utilisateur(request, user_id):
    """Rejette une demande d'inscription en attente : supprime le compte."""
    if request.method == 'POST':
        target = get_object_or_404(CustomUser, id=user_id, status='pending')
        username = target.username
        target.delete()
        Log.objects.create(username=request.user.username, event='USER_REJECT',
                           details=f"Demande d'inscription rejetée : {username}")
        messages.success(request, f"🗑️ Demande de {username} rejetée.")
    return redirect('utilisateurs')


# ==================== ABSENCES ====================

@login_required
def absences_view(request):
    students = Student.objects.all()
    filiere_filter = request.GET.get('filiere', '')
    matricule_filter = request.GET.get('matricule', '')

    absences = Absence.objects.select_related('student').all()
    if filiere_filter:
        absences = absences.filter(student__filiere__iexact=filiere_filter)
    if matricule_filter:
        absences = absences.filter(student__matricule=matricule_filter)

    stats = {}
    for s in students:
        total = Absence.objects.filter(student=s, statut='absent').count()
        justifiees = Absence.objects.filter(student=s, statut='absent', justifiee=True).count()
        stats[s.matricule] = {'total': total, 'justifiees': justifiees, 'injustifiees': total - justifiees}

    if request.method == 'POST':
        matricule = request.POST.get('matricule')
        date_abs = request.POST.get('date')
        matiere = request.POST.get('matiere', '')
        statut = request.POST.get('statut', 'absent')
        justifiee = request.POST.get('justifiee') == 'on'
        commentaire = request.POST.get('commentaire', '')
        student = get_object_or_404(Student, matricule=matricule)
        Absence.objects.create(
            student=student, date=date_abs, matiere=matiere,
            statut=statut, justifiee=justifiee, commentaire=commentaire,
            enregistre_par=request.user.username
        )
        Log.objects.create(username=request.user.username, event='ABSENCE_ADD',
                           details=f"{statut} enregistré pour {student.nom} {student.prenom} le {date_abs}")
        messages.success(request, f'Absence enregistrée pour {student.nom} {student.prenom}.')
        return redirect('absences')

    filieres = Student.objects.values_list('filiere', flat=True).distinct().exclude(filiere__isnull=True).exclude(filiere='')
    return render(request, 'students/absences.html', {
        'students': students, 'absences': absences[:50], 'stats': stats,
        'filieres': filieres, 'filiere_filter': filiere_filter, 'matricule_filter': matricule_filter,
    })


@login_required
def supprimer_absence(request, absence_id):
    if request.method == 'POST':
        absence = get_object_or_404(Absence, id=absence_id)
        absence.delete()
        messages.success(request, 'Absence supprimée.')
    return redirect('absences')


# ==================== EMPLOI DU TEMPS ====================

@login_required
def emploi_du_temps(request):
    filiere_filter = request.GET.get('filiere', '')
    niveau_filter = request.GET.get('niveau', '')

    cours = Cours.objects.all()
    if filiere_filter:
        cours = cours.filter(filiere__iexact=filiere_filter)
    if niveau_filter:
        cours = cours.filter(niveau__iexact=niveau_filter)

    can_manage_schedule = request.user.has_permission('can_manage_schedule')

    if request.method == 'POST':
        if not can_manage_schedule:
            messages.error(request, "⛔ Seul un administrateur peut modifier l'emploi du temps.")
            return redirect(f'emploi-du-temps?filiere={filiere_filter}&niveau={niveau_filter}')
        action = request.POST.get('action')
        if action == 'add':
            Cours.objects.create(
                filiere=request.POST.get('filiere', ''),
                niveau=request.POST.get('niveau', ''),
                matiere=request.POST.get('matiere', ''),
                professeur=request.POST.get('professeur', ''),
                jour=request.POST.get('jour', ''),
                heure_debut=request.POST.get('heure_debut'),
                heure_fin=request.POST.get('heure_fin'),
                salle=request.POST.get('salle', ''),
            )
            messages.success(request, 'Cours ajouté.')
        elif action == 'delete':
            cours_id = request.POST.get('cours_id')
            Cours.objects.filter(id=cours_id).delete()
            messages.success(request, 'Cours supprimé.')
        return redirect(f'emploi-du-temps?filiere={filiere_filter}&niveau={niveau_filter}')

    jours = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi']
    planning = {jour: list(cours.filter(jour=jour)) for jour in jours}
    filieres = Student.objects.values_list('filiere', flat=True).distinct().exclude(filiere__isnull=True).exclude(filiere='')
    niveaux = Student.objects.values_list('niveau', flat=True).distinct().exclude(niveau__isnull=True).exclude(niveau='')

    return render(request, 'students/emploi_du_temps.html', {
        'planning': planning, 'jours': jours, 'cours_list': cours,
        'filieres': filieres, 'niveaux': niveaux,
        'filiere_filter': filiere_filter, 'niveau_filter': niveau_filter,
        'can_manage_schedule': can_manage_schedule,
    })


# ==================== ANNONCES ====================

@login_required
def annonces_view(request):
    current_user = request.user  # déjà un CustomUser (AUTH_USER_MODEL)

    query = request.GET.get('q', '').strip()
    categorie_filter = request.GET.get('categorie', '').strip()

    annonces = Annonce.objects.filter(est_publiee=True)

    if categorie_filter:
        annonces = annonces.filter(categorie=categorie_filter)
    if query:
        annonces = annonces.filter(Q(titre__icontains=query) | Q(contenu__icontains=query))

    annonces = annonces.order_by('-date_publication')

    if request.method == 'POST':
        titre = request.POST.get('titre', '').strip()
        contenu = request.POST.get('contenu', '').strip()
        categorie = request.POST.get('categorie', 'info')
        fichier_joint = request.FILES.get('fichier_joint')

        if fichier_joint:
            extension = fichier_joint.name.rsplit('.', 1)[-1].lower() if '.' in fichier_joint.name else ''
            if extension not in ALLOWED_ATTACHMENT_EXTENSIONS:
                messages.error(
                    request,
                    f"❌ Type de fichier non autorisé. Formats acceptés : "
                    f"{', '.join(ALLOWED_ATTACHMENT_EXTENSIONS)}."
                )
                return redirect('annonces')

        if titre and contenu:
            Annonce.objects.create(
                auteur=current_user,
                titre=titre,
                contenu=contenu,
                categorie=categorie,
                fichier_joint=fichier_joint,
            )
            Log.objects.create(
                username=request.user.username,
                event='ANNONCE_ADD',
                details=f"Nouvelle annonce: {titre}"
            )
            messages.success(request, '✅ Annonce publiée avec succès.')
        else:
            messages.error(request, '❌ Le titre et le contenu sont obligatoires.')
        return redirect('annonces')

    return render(request, 'students/annonces.html', {
        'annonces': annonces,
        'query': query,
        'categorie_filter': categorie_filter,
    })


@login_required
def supprimer_annonce(request, annonce_id):
    if request.method == 'POST':
        annonce = get_object_or_404(Annonce, id=annonce_id)
        # Seul l'auteur ou un admin peut supprimer
        if request.user.role == 'Admin' or annonce.auteur == request.user:
            titre = annonce.titre
            annonce.delete()
            Log.objects.create(
                username=request.user.username,
                event='ANNONCE_DELETE',
                details=f"Annonce supprimée: {titre}"
            )
            messages.success(request, '✅ Annonce supprimée.')
        else:
            messages.error(request, '⛔ Vous n\'avez pas les droits pour supprimer cette annonce.')
    return redirect('annonces')


# ==================== MESSAGERIE INTERNE ====================

@login_required
def messagerie_view(request):
    current_user = request.user  # déjà un CustomUser (AUTH_USER_MODEL)

    messages_recus = Message.objects.filter(destinataire=current_user).order_by('-date_envoi')
    messages_envoyes = Message.objects.filter(expediteur=current_user).order_by('-date_envoi')
    non_lus = messages_recus.filter(lu=False).count()
    utilisateurs = CustomUser.objects.exclude(username=request.user.username).filter(status='active')

    if request.method == 'POST':
        dest_username = request.POST.get('destinataire')
        sujet = request.POST.get('sujet', '').strip()
        contenu = request.POST.get('contenu', '').strip()
        if dest_username and sujet and contenu:
            try:
                destinataire = CustomUser.objects.get(username=dest_username)
                msg = Message.objects.create(
                    expediteur=current_user, destinataire=destinataire,
                    sujet=sujet, contenu=contenu
                )
                notify_user(destinataire.id, {
                    'event': 'new_message',
                    'id': msg.id,
                    'expediteur': current_user.username,
                    'sujet': msg.sujet,
                    'date_envoi': msg.date_envoi.strftime('%d/%m/%Y %H:%M'),
                    'non_lus': Message.objects.filter(destinataire=destinataire, lu=False).count(),
                })
                messages.success(request, f'Message envoyé à {dest_username}.')
            except CustomUser.DoesNotExist:
                messages.error(request, 'Destinataire introuvable.')
        else:
            messages.error(request, 'Tous les champs sont obligatoires.')
        return redirect('messagerie')

    return render(request, 'students/messagerie.html', {
        'messages_recus': messages_recus[:20],
        'messages_envoyes': messages_envoyes[:20],
        'non_lus': non_lus,
        'utilisateurs': utilisateurs,
    })


@login_required
def lire_message(request, message_id):
    current_user = request.user  # déjà un CustomUser (AUTH_USER_MODEL)
    msg = get_object_or_404(Message, id=message_id, destinataire=current_user)
    msg.lu = True
    msg.save()
    return render(request, 'students/lire_message.html', {'msg': msg})


# ==================== IMPORT CSV ====================

@login_required
@permission_required('can_export_data')
def import_csv(request):
    if request.method == 'POST' and request.FILES.get('fichier_csv'):
        fichier = request.FILES['fichier_csv']
        decoded = fichier.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(decoded))
        ajoutes = erreurs = 0
        for row in reader:
            try:
                matricule = str(row.get('Matricule', '') or row.get('matricule', '')).strip()
                nom = str(row.get('Nom', '') or row.get('nom', '')).strip().capitalize()
                prenom = str(row.get('Prénom', '') or row.get('prenom', '')).strip().capitalize()
                age = int(row.get('Âge', 0) or row.get('age', 0) or 18)
                filiere = str(row.get('Filière', '') or row.get('filiere', '')).strip().upper()
                niveau = str(row.get('Niveau', '') or row.get('niveau', '')).strip()
                email = str(row.get('Email', '') or row.get('email', '')).strip()
                telephone = str(row.get('Téléphone', '') or row.get('telephone', '')).strip()
                statut = str(row.get('Statut', 'Actif') or 'Actif').strip()
                notes_str = str(row.get('Notes', '') or '').strip()
                notes_vals = parse_notes(notes_str) if notes_str else []

                if not matricule or not nom or not prenom:
                    erreurs += 1
                    continue
                if Student.objects.filter(matricule=matricule).exists():
                    erreurs += 1
                    continue

                student = Student.objects.create(
                    matricule=matricule, nom=nom, prenom=prenom, age=age,
                    filiere=filiere, niveau=niveau, email=email or None,
                    telephone=telephone or None, statut=statut,
                )
                # Créer les notes via le modèle dédié
                for val in notes_vals:
                    Note.objects.create(student=student, valeur=val)
                ajoutes += 1
            except Exception:
                erreurs += 1

        Log.objects.create(username=request.user.username, event='IMPORT_CSV',
                           details=f'{ajoutes} importé(s), {erreurs} erreur(s)')
        messages.success(request, f'✅ {ajoutes} étudiant(s) importé(s). {erreurs} ligne(s) ignorée(s).')
        return redirect('index')

    return render(request, 'students/import_csv.html')


# ==================== CLASSEMENT ====================

@login_required
def classement_view(request):
    filiere_filter = request.GET.get('filiere', '')
    niveau_filter = request.GET.get('niveau', '')
    students = Student.objects.all().prefetch_related('notes')
    if filiere_filter:
        students = students.filter(filiere__iexact=filiere_filter)
    if niveau_filter:
        students = students.filter(niveau__iexact=niveau_filter)

    classement = []
    for s in students:
        notes_list = [float(n.valeur) for n in s.notes.all()]
        moy = calculate_average(notes_list)
        classement.append({'student': s, 'moyenne': moy, 'appreciation': get_appreciation(moy)})
    classement.sort(key=lambda x: x['moyenne'], reverse=True)
    for i, item in enumerate(classement, 1):
        item['rang'] = i

    filieres = Student.objects.values_list('filiere', flat=True).distinct().exclude(filiere__isnull=True).exclude(filiere='')
    niveaux = Student.objects.values_list('niveau', flat=True).distinct().exclude(niveau__isnull=True).exclude(niveau='')

    return render(request, 'students/classement.html', {
        'classement': classement,
        'filieres': filieres, 'niveaux': niveaux,
        'filiere_filter': filiere_filter, 'niveau_filter': niveau_filter,
    })


# ==================== CARTE ÉTUDIANT PRÉVISUALISATION ====================

@login_required
@permission_required('can_generate_documents')
def carte_preview(request, matricule):
    student = get_object_or_404(Student, matricule=matricule)
    notes_list = [float(n.valeur) for n in student.notes.all()]
    moyenne = calculate_average(notes_list)
    appreciation = get_appreciation(moyenne)
    from django.utils import timezone
    year = timezone.now().year
    return render(request, 'students/carte_preview.html', {
        'student': student,
        'notes_list': notes_list,
        'moyenne': moyenne,
        'appreciation': appreciation,
        'year': year,
    })


# ==================== CARTE ÉTUDIANT PDF ====================

@login_required
@permission_required('can_generate_documents')
def carte_etudiant_pdf(request, matricule):
    if not REPORTLAB_AVAILABLE:
        messages.error(request, 'ReportLab n\'est pas installé.')
        return redirect('index')

    student = get_object_or_404(Student, matricule=matricule)
    notes_list = [float(n.valeur) for n in student.notes.all()]
    moyenne = calculate_average(notes_list)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="carte_{matricule}.pdf"'

    from reportlab.lib.pagesizes import landscape
    from reportlab.platypus import HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    card_w, card_h = 9.5 * cm, 6 * cm
    doc = SimpleDocTemplate(response, pagesize=(card_w, card_h),
                            topMargin=0.3*cm, bottomMargin=0.3*cm,
                            leftMargin=0.3*cm, rightMargin=0.3*cm)
    elements = []
    styles = getSampleStyleSheet()

    header_style = ParagraphStyle('H', parent=styles['Normal'], alignment=1,
                                   fontSize=8, textColor=colors.white, fontName='Helvetica-Bold')
    body_style = ParagraphStyle('B', parent=styles['Normal'], fontSize=7,
                                 textColor=colors.HexColor('#1e293b'), leading=11)
    label_style = ParagraphStyle('L', parent=styles['Normal'], fontSize=6,
                                  textColor=colors.HexColor('#64748b'), leading=9)

    # En-tête coloré
    header_data = [['🎓 STUDENT MANAGER PRO', 'CARTE ÉTUDIANT']]
    header_table = Table(header_data, colWidths=[5.5*cm, 3.4*cm])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 4))

    # Corps de la carte
    body_data = [
        [Paragraph(f'<b>{student.nom.upper()} {student.prenom}</b>', body_style), ''],
        [Paragraph(f'Matricule : <b>{student.matricule}</b>', body_style), ''],
        [Paragraph(f'Filière : {student.filiere or "-"}  |  Niveau : {student.niveau or "-"}', label_style), ''],
        [Paragraph(f'Moyenne : <b>{moyenne:.2f}/20</b>  |  {get_appreciation(moyenne)}', label_style), ''],
        [Paragraph(f'Statut : {student.statut}', label_style), ''],
    ]
    body_table = Table(body_data, colWidths=[7*cm, 2.2*cm])
    body_table.setStyle(TableStyle([
        ('PADDING', (0, 0), (-1, -1), 3),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(body_table)
    elements.append(Spacer(1, 4))

    # Pied de carte
    footer_data = [[
        Paragraph(f'Année {timezone.now().year}-{timezone.now().year+1}', label_style),
        Paragraph('Signature : ___________', label_style)
    ]]
    footer_table = Table(footer_data, colWidths=[4.5*cm, 4.4*cm])
    footer_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f1f5f9')),
        ('PADDING', (0, 0), (-1, -1), 4),
        ('LINEABOVE', (0, 0), (-1, 0), 0.5, colors.HexColor('#4f46e5')),
    ]))
    elements.append(footer_table)

    doc.build(elements)
    Log.objects.create(username=request.user.username, event='CARTE_PDF',
                       details=f'Carte générée pour {student.nom} {student.prenom} ({matricule})')
    return response


# ==================== RAPPELS EMAIL ====================

@login_required
@permission_required('can_view_finances')
def envoyer_rappels_email(request):
    from django.core.mail import send_mail
    from django.conf import settings as django_settings
    if request.method != 'POST':
        return redirect('paiements')
    etudiants_en_retard = [s for s in Student.objects.all() if s.statut_paiement() == 'En retard' and s.email]
    envoyes = erreurs = 0
    for student in etudiants_en_retard:
        solde = float(student.solde_restant())
        sujet = f'[Student Manager PRO] Rappel de paiement - {student.nom} {student.prenom}'
        corps = (
            f'Bonjour {student.prenom} {student.nom},\n\n'
            f'Votre dossier présente un solde impayé de {solde:,.0f} FCFA.\n'
            f'Matricule : {student.matricule} | Filière : {student.filiere or "-"} | Niveau : {student.niveau or "-"}\n\n'
            f'Merci de régulariser votre situation.\n\nCordialement,\nL\'Administration'
        )
        try:
            send_mail(sujet, corps, django_settings.DEFAULT_FROM_EMAIL, [student.email], fail_silently=False)
            envoyes += 1
        except Exception:
            erreurs += 1
    Log.objects.create(username=request.user.username, event='EMAIL_RAPPELS',
                       details=f'{envoyes} rappel(s) envoyé(s), {erreurs} échec(s)')
    if envoyes:
        messages.success(request, f'✅ {envoyes} rappel(s) envoyé(s).')
    if erreurs:
        messages.warning(request, f'⚠️ {erreurs} envoi(s) échoué(s). Vérifiez la config email dans .env.')
    if not etudiants_en_retard:
        messages.info(request, 'Aucun étudiant en retard avec email enregistré.')
    return redirect('paiements')


# ==================== BULLETIN PRÉVISUALISATION ====================

@login_required
@permission_required('can_generate_documents')
def bulletins_list(request):
    """Page listant tous les étudiants avec accès à leurs bulletins"""
    students = Student.objects.all().order_by('-matricule').prefetch_related('notes')
    students_data = []
    for s in students:
        notes_list = [float(n.valeur) for n in s.notes.all()]
        moyenne = calculate_average(notes_list)
        students_data.append({
            'student': s,
            'moyenne': moyenne,
        })
    return render(request, 'students/bulletins_list.html', {
        'students_data': students_data,
    })


@login_required
@permission_required('can_generate_documents')
def bulletin_preview(request, matricule):
    student = get_object_or_404(Student, matricule=matricule)
    notes_list = [float(n.valeur) for n in student.notes.all()]
    moyenne = calculate_average(notes_list)
    appreciation = get_appreciation(moyenne)
    return render(request, 'students/bulletin_preview.html', {
        'student': student,
        'notes_list': notes_list,
        'moyenne': moyenne,
        'appreciation': appreciation,
    })


# ==================== BULLETIN PDF ====================

@login_required
@permission_required('can_generate_documents')
def bulletin_pdf(request, matricule):
    if not REPORTLAB_AVAILABLE:
        messages.error(request, 'ReportLab n\'est pas installé.')
        return redirect('index')
    student = get_object_or_404(Student, matricule=matricule)
    notes_list = [float(n.valeur) for n in student.notes.all()]
    moyenne = calculate_average(notes_list)
    appreciation = get_appreciation(moyenne)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="bulletin_{matricule}.pdf"'
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('T', parent=styles['Heading1'], alignment=1, textColor=colors.HexColor('#4f46e5'), spaceAfter=6)
    section_style = ParagraphStyle('S', parent=styles['Heading2'], textColor=colors.HexColor('#1e293b'), spaceBefore=16, spaceAfter=8)
    elements.append(Paragraph('STUDENT MANAGER PRO', title_style))
    elements.append(Paragraph('BULLETIN DE NOTES', title_style))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph('Informations de l\'étudiant', section_style))
    info_data = [
        ['Matricule', str(student.matricule), 'Statut', student.statut],
        ['Nom', student.nom, 'Prénom', student.prenom],
        ['Filière', student.filiere or '-', 'Niveau', student.niveau or '-'],
        ['Âge', str(student.age), 'Email', student.email or '-'],
    ]
    info_table = Table(info_data, colWidths=[3.5*cm, 7*cm, 3.5*cm, 7*cm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f5f9')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#f1f5f9')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('PADDING', (0, 0), (-1, -1), 8),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 16))
    elements.append(Paragraph('Résultats Académiques', section_style))
    if notes_list:
        notes_data = [['N°', 'Note', 'Appréciation']]
        for i, note_val in enumerate(notes_list, 1):
            notes_data.append([str(i), f'{note_val:.2f} / 20', get_appreciation(note_val)])
        notes_data.append(['', 'MOYENNE', f'{moyenne:.2f} / 20'])
        notes_table = Table(notes_data, colWidths=[2*cm, 6*cm, 13*cm])
        notes_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4f46e5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e0e7ff')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(notes_table)
    else:
        elements.append(Paragraph('Aucune note enregistrée.', styles['Normal']))
    elements.append(Spacer(1, 20))
    recap_data = [
        ['Moyenne Générale', f'{moyenne:.2f} / 20'],
        ['Mention', appreciation],
        ['Décision', 'ADMIS(E)' if moyenne >= 10 else 'AJOURNÉ(E)'],
    ]
    recap_table = Table(recap_data, colWidths=[8*cm, 13*cm])
    recap_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('PADDING', (0, 0), (-1, -1), 10),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
    ]))
    elements.append(recap_table)
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(
        f'Généré le {timezone.now().strftime("%d/%m/%Y à %H:%M")} | Student Manager PRO',
        ParagraphStyle('Footer', parent=styles['Normal'], alignment=1, textColor=colors.HexColor('#94a3b8'), fontSize=8)
    ))
    doc.build(elements)
    Log.objects.create(username=request.user.username, event='BULLETIN_PDF',
                       details=f'Bulletin généré pour {student.nom} {student.prenom} ({matricule})')
    return response


# ==================== STATISTIQUES AVANCÉES ====================

@login_required
def stats_avancees(request):
    students = list(Student.objects.all().prefetch_related('notes'))
    notes_all = []
    for s in students:
        s_notes = [float(n.valeur) for n in s.notes.all()]
        if s_notes:
            notes_all.append(calculate_average(s_notes))
    mentions = {'Excellent': 0, 'Très_bien': 0, 'Bien': 0, 'Assez_bien': 0, 'Passable': 0, 'Insuffisant': 0}
    for m in notes_all:
        if m >= 18: mentions['Excellent'] += 1
        elif m >= 16: mentions['Très_bien'] += 1
        elif m >= 14: mentions['Bien'] += 1
        elif m >= 12: mentions['Assez_bien'] += 1
        elif m >= 10: mentions['Passable'] += 1
        else: mentions['Insuffisant'] += 1
    filieres_data = {}
    for s in students:
        f = s.filiere or 'Non définie'
        if f not in filieres_data: filieres_data[f] = []
        s_notes = [float(n.valeur) for n in s.notes.all()]
        if s_notes: filieres_data[f].append(calculate_average(s_notes))
    moy_par_filiere = {f: round(sum(v)/len(v), 2) for f, v in filieres_data.items() if v}
    niveaux_data = {}
    for s in students:
        n = s.niveau or 'Non défini'
        if n not in niveaux_data: niveaux_data[n] = []
        s_notes = [float(n.valeur) for n in s.notes.all()]
        if s_notes: niveaux_data[n].append(calculate_average(s_notes))
    moy_par_niveau = {n: round(sum(v)/len(v), 2) for n, v in niveaux_data.items() if v}
    # Données préparées ici, graphiques rendus en Chart.js côté navigateur
    # (le rendu serveur via Matplotlib était trop coûteux en CPU sur l'hébergement gratuit)
    from collections import Counter
    charts = {}

    # 1. Camembert des mentions
    mentions_labels = [k for k, v in mentions.items() if v > 0]
    mentions_values = [v for k, v in mentions.items() if v > 0]
    if mentions_values:
        charts['mentions'] = {'labels': mentions_labels, 'values': mentions_values}

    # 2. Barres des moyennes par filière
    if moy_par_filiere:
        charts['filieres'] = {'labels': list(moy_par_filiere.keys()), 'values': list(moy_par_filiere.values())}

    # 3. Histogramme distribution des moyennes (10 tranches de 0 à 20)
    if notes_all:
        bins = [0] * 10
        for m in notes_all:
            bins[min(int(m // 2), 9)] += 1
        charts['distribution'] = {
            'labels': [f'{i*2}-{i*2+2}' for i in range(10)],
            'values': bins,
        }

    # 4. Barres des moyennes par niveau
    if moy_par_niveau:
        charts['niveaux'] = {'labels': list(moy_par_niveau.keys()), 'values': list(moy_par_niveau.values())}

    # 5. Taux de réussite par filière
    if filieres_data:
        taux_par_filiere = {f: round(sum(1 for m in v if m >= 10) / len(v) * 100, 1)
                             for f, v in filieres_data.items() if v}
        if taux_par_filiere:
            charts['taux_reussite_filiere'] = {
                'labels': list(taux_par_filiere.keys()), 'values': list(taux_par_filiere.values()),
            }

    # 6. Répartition par statut
    actifs = sum(1 for s in students if s.statut == 'Actif')
    suspendus = sum(1 for s in students if s.statut == 'Suspendu')
    statut_pairs = [(l, v) for l, v in [('Actifs', actifs), ('Suspendus', suspendus)] if v > 0]
    if statut_pairs:
        charts['statut'] = {'labels': [p[0] for p in statut_pairs], 'values': [p[1] for p in statut_pairs]}

    # 7. Répartition par filière
    filiere_counts = Counter(s.filiere or 'Non définie' for s in students)
    if filiere_counts:
        charts['repartition_filieres'] = {
            'labels': list(filiere_counts.keys()), 'values': list(filiere_counts.values()),
        }

    # 8. Évolution des inscriptions par mois
    inscriptions_par_mois = Counter()
    for s in students:
        if s.date_ajout:
            inscriptions_par_mois[s.date_ajout.strftime('%Y-%m')] += 1
    if inscriptions_par_mois:
        sorted_mois = sorted(inscriptions_par_mois.items())
        charts['inscriptions'] = {
            'labels': [m[0][-2:] + '/' + m[0][:4] for m in sorted_mois],
            'values': [m[1] for m in sorted_mois],
        }

    taux_reussite = round(sum(1 for m in notes_all if m >= 10) / len(notes_all) * 100, 1) if notes_all else 0
    return render(request, 'students/stats.html', {
        'mentions': mentions, 'moy_par_filiere': moy_par_filiere, 'moy_par_niveau': moy_par_niveau,
        'taux_reussite': taux_reussite, 'total': len(students), 'charts': charts,
    })


# ==================== RAPPORTS FINANCIERS ====================

@login_required
@permission_required('can_view_finances')
def rapports_financiers(request):
    """Page dédiée aux rapports et graphiques financiers"""
    from collections import Counter

    # Une seule requête agrégée (total payé par étudiant) plutôt que
    # total_paye()/statut_paiement() en boucle Python (N+1 queries)
    students_list = list(Student.with_totals())

    total_scolarite = sum(float(s.frais_scolarite) for s in students_list)
    total_encaisse = sum(float(s.total_paye_annot) for s in students_list)
    total_dettes = total_scolarite - total_encaisse
    statuts = {s.matricule: compute_statut_paiement(s.total_paye_annot, s.frais_scolarite) for s in students_list}
    a_jour = sum(1 for v in statuts.values() if v == 'À jour')
    en_retard = sum(1 for v in statuts.values() if v == 'En retard')
    partiel = sum(1 for v in statuts.values() if v == 'Partiel')
    alertes = [s for s in students_list if statuts[s.matricule] == 'En retard'][:10]

    # Données préparées ici, graphiques rendus en Chart.js côté navigateur
    charts = {}

    # 1. Évolution des paiements sur 12 mois
    raw_paiements = Payment.objects.filter(
        date_paiement__gte=timezone.now() - timedelta(days=365)
    ).values_list('date_paiement', 'montant').order_by('date_paiement')

    mois_dict = {}
    for dp, montant in raw_paiements:
        if dp:
            key = dp.strftime('%Y-%m') if hasattr(dp, 'strftime') else str(dp)[:7]
            mois_dict[key] = mois_dict.get(key, 0) + float(montant)

    paiements_list = sorted(mois_dict.items())
    if paiements_list:
        charts['paiements'] = {
            'labels': [m[-2:] for m, _ in paiements_list],
            'values': [round(v, 2) for _, v in paiements_list],
        }

    # 2. Taux d'impayés par filière (agrégé en Python à partir de with_totals(), pas de requête par filière)
    filieres_totals = {}
    for s in students_list:
        if not s.filiere:
            continue
        entry = filieres_totals.setdefault(s.filiere, {'du': Decimal('0.00'), 'paye': Decimal('0.00')})
        entry['du'] += s.frais_scolarite
        entry['paye'] += s.total_paye_annot

    noms_filieres, taux_impayes = [], []
    for filiere, totaux in filieres_totals.items():
        if totaux['du'] > 0:
            noms_filieres.append(filiere)
            taux_impayes.append(round(float((totaux['du'] - totaux['paye']) / totaux['du'] * 100), 1))

    if noms_filieres:
        charts['impayes'] = {'labels': noms_filieres, 'values': taux_impayes}

    # 3. Répartition des statuts de paiement
    statut_pairs = [(l, v) for l, v in [('À jour', a_jour), ('Partiel', partiel), ('En retard', en_retard)] if v > 0]
    if statut_pairs:
        charts['statuts_paiement'] = {'labels': [p[0] for p in statut_pairs], 'values': [p[1] for p in statut_pairs]}

    # 4. Montants par filière (scolarité due vs encaissée)
    filieres_montants = {}
    for s in students_list:
        f = s.filiere or 'Non définie'
        entry = filieres_montants.setdefault(f, {'scolarite': 0.0, 'paye': 0.0})
        entry['scolarite'] += float(s.frais_scolarite)
        entry['paye'] += float(s.total_paye_annot)
    if filieres_montants:
        charts['montants_filieres'] = {
            'labels': list(filieres_montants.keys()),
            'scolarite': [round(v['scolarite'], 2) for v in filieres_montants.values()],
            'paye': [round(v['paye'], 2) for v in filieres_montants.values()],
        }

    # 5. Évolution des encaissements par mois
    paiements_par_mois = Counter()
    for p in Payment.objects.all():
        if p.date_paiement:
            paiements_par_mois[p.date_paiement.strftime('%Y-%m')] += float(p.montant)
    if paiements_par_mois:
        sorted_p = sorted(paiements_par_mois.items())
        charts['encaissements'] = {
            'labels': [m[-2:] + '/' + m[:4] for m, _ in sorted_p],
            'values': [round(v, 2) for _, v in sorted_p],
        }

    return render(request, 'students/rapports_financiers.html', {
        'total_scolarite': total_scolarite,
        'total_encaisse': total_encaisse,
        'total_dettes': total_dettes,
        'a_jour': a_jour,
        'en_retard': en_retard,
        'partiel': partiel,
        'alertes': alertes,
        'charts': charts,
    })


# ==================== NOTIFICATIONS ====================

@login_required
def notifications_view(request):
    """Liste toutes les notifications de l'utilisateur connecté"""
    notifications = Notification.objects.filter(destinataire=request.user)
    return render(request, 'students/notifications.html', {
        'notifications': notifications,
    })


@login_required
def notification_mark_read(request, notification_id):
    """Marque une notification comme lue"""
    if request.method == 'POST':
        notif = get_object_or_404(Notification, id=notification_id, destinataire=request.user)
        notif.lu = True
        notif.save()
    return redirect('notifications')


@login_required
def notifications_mark_all_read(request):
    """Marque toutes les notifications de l'utilisateur comme lues"""
    if request.method == 'POST':
        Notification.objects.filter(destinataire=request.user, lu=False).update(lu=True)
    return redirect('notifications')
